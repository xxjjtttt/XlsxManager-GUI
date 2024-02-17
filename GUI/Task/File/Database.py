"""
Author: xxjjtttt
Date: 2024-01-22 17:06:27
LastEditTime: 2024-01-22 17:55:23
LastEditors: xxjjtttt
Description: https://github.com/xxjjtttt
FilePath: \\python\\xlsx master\\File\\Database.py
屎山代码什么时候重构
"""

# python -m pip install upgrade pip
# pip install openpyxl==3.1.2
from openpyxl import load_workbook as load
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class Database:
    file: Workbook
    path: str
    working_sheet: Worksheet
    working_row: int
    active_sheet: list[int]

    def __init__(self, path: str, active_sheet: list[int]) -> None:
        self.path = path
        self.active_sheet = active_sheet
        self.file = load(path)
        self.working_sheet = self.file[self.file.sheetnames[0]]
        self.working_row = 2

    def get_data_list(self) -> list:
        data_list: list = []
        max_column: int = self.working_sheet.max_column
        for column in range(1, max_column + 1):
            temp_data = self.working_sheet.cell(self.working_row, column).value
            if temp_data is not None:
                data_list.append(temp_data)
            else:
                break
        self.working_row += 1
        return data_list

    def get_sheetnames(self) -> list[str]:
        if not self.active_sheet:
            return self.file.sheetnames
        else:
            sheetnames: list[str] = self.file.sheetnames
            active_sheetnames: list[str] = []
            for index in self.active_sheet:
                active_sheetnames.append(sheetnames[index - 1])
            return active_sheetnames

    def change_working_sheet(self, sheetname: str):
        self.working_sheet = self.file[sheetname]
        self.working_row = 2

    def get_max(self) -> list:
        return [self.working_sheet.max_row - 1, self.working_sheet.max_column]

    def close(self) -> None:
        self.file.close()
