"""
Author: xxjjtttt
Date: 2024-01-22 16:53:03
LastEditTime: 2024-01-22 22:31:44
LastEditors: xxjjtttt
Description: https://github.com/xxjjtttt
FilePath: \\python\\xlsx master\\File\\Sourcefile.py
屎山代码什么时候重构
"""

# python -m pip install upgrade pip
# pip install openpyxl==3.1.2

from typing import Any

from openpyxl import load_workbook as load
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class Sourcefile:
    file: Workbook
    path: str
    working_sheet: Worksheet
    working_row: int
    key_column: int
    index_column: int
    have_header: bool
    active_sheet: list[int]

    def __init__(self, path: str, key_column: int, index_column: int,
                 have_header: bool, active_sheet: list[int]) -> None:
        self.path = path
        self.file = load(path)
        self.key_column = key_column
        self.index_column = index_column
        self.have_header = have_header
        self.working_sheet = self.file[self.file.sheetnames[0]]
        self.reset_working_row()
        self.active_sheet = active_sheet

    def get_key_value(self) -> Any:
        return self.working_sheet.cell(self.working_row, self.key_column).value

    def get_source_list(self) -> list:
        source_list: list = []
        max_column: int = self.working_sheet.max_column
        for column in range(1, max_column + 1):
            temp_data = self.working_sheet.cell(self.working_row, column).value
            source_list.append(temp_data)
        self.tag_row()
        return source_list

    def tag_row(self) -> None:
        self.working_sheet.cell(self.working_row, self.key_column).value = "."

    def next_working_row(self) -> None:
        self.working_row += 1

    def reset_working_row(self) -> None:
        if self.have_header:
            self.working_row = 2
        else:
            self.working_row = 1

    def get_max(self) -> list:
        if self.have_header:
            return [self.working_sheet.max_row + 1, self.working_sheet.max_column]
        else:
            return [self.working_sheet.max_row + 1, self.working_sheet.max_column]

    def get_sheetnames(self) -> list[str]:
        if self.active_sheet:
            return self.file.sheetnames
        else:
            sheetnames: list[str] = self.file.sheetnames
            active_sheetnames: list[str] = []
            for index in self.active_sheet:
                active_sheetnames.append(sheetnames[index - 1])
            return active_sheetnames

    def change_working_sheet(self, sheetname: str):
        self.reset_working_row()
        self.working_sheet = self.file[sheetname]

    def get_header_list(self) -> list:
        header_list: list = []
        if self.have_header:
            for column in range(self.working_sheet.max_column):
                header_list.append(self.working_sheet.cell(1, column + 1).value)
        return header_list

    def close(self) -> None:
        self.file.close()

    def throw_rubbish(self) -> None:
        trash = Workbook()
        trash.active.title = "trash"
        trash_sheet = trash["trash"]
        for sheetname in self.get_sheetnames():
            self.change_working_sheet(sheetname)
            for row in range(self.get_max()[0]):
                if self.get_key_value() != ".":
                    trash_sheet.append(self.get_source_list())
                self.next_working_row()
        trash.save("trash.xlsx")
