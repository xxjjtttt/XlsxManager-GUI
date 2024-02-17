"""
Author: xxjjtttt
Date: 2024-01-22 17:06:02
LastEditTime: 2024-01-22 22:24:56
LastEditors: xxjjtttt
Description: https://github.com/xxjjtttt
FilePath: \\python\\xlsx master\\File\\Targetfile.py
屎山代码什么时候重构
"""

# python -m pip install upgrade pip
# pip install openpyxl==3.1.2

# from openpyxl import load_workbook as load
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class Targetfile:
    file: Workbook
    path: str
    working_sheet: Worksheet
    working_row: int
    key_column: int
    index_column: int
    have_header: bool

    def __init__(self, path: str, key_column: int, index_column: int, have_header: bool, sheetnames: list[str]) -> None:
        if path == "":
            self.path = "Untitled.xlsx"
        else:
            self.path = path
        self.file = Workbook()
        self.file.active.title = sheetnames[0]
        for index in range(1, len(sheetnames)):
            self.file.create_sheet(sheetnames[index])
        self.key_column = key_column
        self.index_column = index_column
        self.have_header = have_header
        self.working_sheet = self.file[sheetnames[0]]
        self.__reset_working_row()
        self.save()

    def __reset_working_row(self) -> None:
        if self.have_header:
            self.working_row = 2
        else:
            self.working_row = 1

    def set_header(self, header_list: list) -> None:
        for sheetname in self.file.sheetnames:
            self.file[sheetname].append(header_list)

    def add_row(self, source_list: list) -> None:
        self.working_sheet.append(source_list)

    def get_sheetnames(self) -> list[str]:
        return self.file.sheetnames

    def change_working_sheet(self, sheetname: str):
        self.working_sheet = self.file[sheetname]
        self.__reset_working_row()

    def save(self) -> None:
        self.file.save(self.path)

    def close(self) -> None:
        self.file.close()
